# ============================================================================
# 新星肉羊育种系统 - 数据导入导出服务
# NovaBreed Sheep System - Data Import/Export Service
# ============================================================================

import io
import csv
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from fastapi import UploadFile, HTTPException

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


class DataImportExportService:
    """数据导入导出服务"""
    
    # 实体配置
    ENTITY_CONFIGS = {
        'farms': {
            'model': 'Farm',
            'required_columns': ['code', 'name', 'farm_type', 'capacity'],
            'optional_columns': ['address', 'contact_person', 'contact_phone'],
            'validators': {
                'code': lambda x: len(str(x)) <= 50,
                'name': lambda x: len(str(x)) <= 100,
                'farm_type': lambda x: x in ['breeding', 'commercial', 'research'],
                'capacity': lambda x: int(x) > 0,
            }
        },
        'animals': {
            'model': 'Animal',
            'required_columns': ['animal_id', 'farm_id', 'birth_date', 'gender'],
            'optional_columns': ['sire_id', 'dam_id', 'breed', 'status'],
            'validators': {
                'gender': lambda x: x in ['male', 'female', 'M', 'F'],
            }
        },
        'phenotypes': {
            'model': 'Phenotype',
            'required_columns': ['animal_id', 'trait_code', 'value', 'measure_date'],
            'optional_columns': ['age_days', 'notes'],
            'validators': {
                'value': lambda x: isinstance(float(x), float),
            }
        },
        'genotypes': {
            'model': 'Genotype',
            'required_columns': ['animal_id', 'snp_id', 'genotype'],
            'optional_columns': [],
            'validators': {
                'genotype': lambda x: x in ['AA', 'AB', 'BA', 'BB', '0', '1', '2'],
            }
        },
    }
    
    async def preview_file(
        self, 
        file: UploadFile, 
        entity_type: str, 
        max_rows: int = 100
    ) -> Dict[str, Any]:
        """
        预览导入文件内容
        
        Args:
            file: 上传的文件
            entity_type: 实体类型
            max_rows: 最大预览行数
            
        Returns:
            预览数据和验证结果
        """
        if entity_type not in self.ENTITY_CONFIGS:
            raise HTTPException(status_code=400, detail=f"不支持的实体类型: {entity_type}")
        
        config = self.ENTITY_CONFIGS[entity_type]
        
        # 读取文件
        content = await file.read()
        await file.seek(0)
        
        try:
            if file.filename.endswith('.xlsx'):
                df = pd.read_excel(io.BytesIO(content), nrows=max_rows)
            elif file.filename.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(content), nrows=max_rows)
            else:
                raise HTTPException(status_code=400, detail="不支持的文件格式")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")
        
        # 检查必填列
        missing_columns = [col for col in config['required_columns'] if col not in df.columns]
        
        # 验证数据
        errors = []
        for idx, row in df.iterrows():
            for col, validator in config.get('validators', {}).items():
                if col in df.columns:
                    try:
                        if not validator(row[col]):
                            errors.append({
                                'row': idx + 2,  # Excel行号从1开始，加表头
                                'column': col,
                                'error': f'值 "{row[col]}" 不符合验证规则'
                            })
                    except Exception as e:
                        errors.append({
                            'row': idx + 2,
                            'column': col,
                            'error': str(e)
                        })
        
        return {
            'data': df.to_dict('records'),
            'total_rows': len(df),
            'columns': list(df.columns),
            'missing_columns': missing_columns,
            'errors': errors[:50],  # 最多返回50个错误
            'is_valid': len(missing_columns) == 0 and len(errors) == 0
        }
    
    async def import_data(
        self, 
        file: UploadFile, 
        entity_type: str,
        db_session: Any
    ) -> Dict[str, Any]:
        """
        导入数据到数据库
        
        Args:
            file: 上传的文件
            entity_type: 实体类型
            db_session: 数据库会话
            
        Returns:
            导入结果统计
        """
        preview = await self.preview_file(file, entity_type, max_rows=None)
        
        if not preview['is_valid']:
            return {
                'success': 0,
                'failed': len(preview['data']),
                'errors': preview['errors']
            }
        
        success_count = 0
        failed_count = 0
        errors = []
        
        for idx, record in enumerate(preview['data']):
            try:
                # 这里实际应该插入数据库
                # await self._insert_record(entity_type, record, db_session)
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append({
                    'row': idx + 2,
                    'error': str(e)
                })
        
        logger.info(f"导入完成: {entity_type}, 成功: {success_count}, 失败: {failed_count}")
        
        return {
            'success': success_count,
            'failed': failed_count,
            'errors': errors[:50]
        }
    
    async def export_data(
        self, 
        entity_type: str, 
        format: str, 
        filters: Optional[Dict] = None,
        db_session: Any = None
    ) -> Tuple[bytes, str]:
        """
        导出数据
        
        Args:
            entity_type: 实体类型
            format: 导出格式 (xlsx, csv)
            filters: 过滤条件
            db_session: 数据库会话
            
        Returns:
            (文件内容, 文件名)
        """
        # 模拟数据 - 实际应该从数据库查询
        sample_data = [
            {'id': 1, 'code': 'FARM001', 'name': '示范羊场1', 'farm_type': 'breeding', 'capacity': 1000},
            {'id': 2, 'code': 'FARM002', 'name': '示范羊场2', 'farm_type': 'commercial', 'capacity': 2000},
        ]
        
        df = pd.DataFrame(sample_data)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'xlsx':
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='数据')
                
                # 美化Excel
                ws = writer.sheets['数据']
                header_fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
            
            output.seek(0)
            return output.getvalue(), f'{entity_type}_export_{timestamp}.xlsx'
        
        elif format == 'csv':
            output = io.StringIO()
            df.to_csv(output, index=False)
            return output.getvalue().encode('utf-8-sig'), f'{entity_type}_export_{timestamp}.csv'
        
        else:
            raise HTTPException(status_code=400, detail=f"不支持的导出格式: {format}")
    
    def generate_template(self, entity_type: str) -> Tuple[bytes, str]:
        """
        生成导入模板
        
        Args:
            entity_type: 实体类型
            
        Returns:
            (文件内容, 文件名)
        """
        if entity_type not in self.ENTITY_CONFIGS:
            raise HTTPException(status_code=400, detail=f"不支持的实体类型: {entity_type}")
        
        config = self.ENTITY_CONFIGS[entity_type]
        columns = config['required_columns'] + config.get('optional_columns', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = '数据模板'
        
        # 添加表头
        header_fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 15
        
        # 添加必填标识行
        required_row = []
        for col in columns:
            if col in config['required_columns']:
                required_row.append('必填')
            else:
                required_row.append('选填')
        
        for col_idx, value in enumerate(required_row, 1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            if value == '必填':
                cell.fill = PatternFill(start_color='FFCCC7', end_color='FFCCC7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # 添加示例数据
        if entity_type == 'farms':
            ws.cell(row=3, column=1, value='FARM001')
            ws.cell(row=3, column=2, value='示范羊场')
            ws.cell(row=3, column=3, value='breeding')
            ws.cell(row=3, column=4, value=1000)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), f'{entity_type}_template.xlsx'


# 创建服务实例
data_io_service = DataImportExportService()
